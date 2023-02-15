import openpyxl

def getRowCount(path, sheetName):
    wb=openpyxl.load_workbook(path)
    sh=wb[sheetName]
    return(sh.max_row)

def getColumnCount(path, sheetName):
    wb= openpyxl.load_workbook(path)
    sh=wb[sheetName]
    return(sh.max_column)

def readData(path, sheetName, rownum, columnnum):
    wb=openpyxl.load_workbook(path)
    sh=wb[sheetName]
    return sh.cell(row=rownum, column=columnnum).value

def writeData(path, sheetName, rownum, columnnum, data):
    wb= openpyxl.load_workbook(path)
    sh=wb[sheetName]
    sh.cell(row=rownum, column=columnnum).value=data
    wb.save(path)